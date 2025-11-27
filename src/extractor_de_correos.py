import os
import re
import win32com.client
import pandas as pd
import unicodedata
import time

from datetime import datetime
from datetime import timedelta

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from colorama import init, Fore

def limpiar_acortar_remitentes(texto):
    if not texto:
        return ""
    texto = str(texto).strip()
    
    # Reemplazar caracteres inválidos + comillas Unicode
    texto = re.sub(
        r'[\\/:*?"<>|\r\n\t“”‘’´`]',
        "_",
        texto
    )
    
    # Normalizar espacios
    texto = re.sub(r'\s+', " ", texto)
    texto = texto.strip()

    # Limitar longitud
    if len(texto) > 45:
        texto = texto[:45]

    return f"{texto}"
    
def limpiar_texto(nombre):
    if not nombre:
        return ""
    nombre = str(nombre).strip()
    
    # Detectar extensión válida (opcional)
    m = re.match(r"(.+)\.([A-Za-z0-9]{1,5})$", nombre)
    if m:
        base, ext = m.group(1), "." + m.group(2).lower()
    else:
        base, ext = nombre, ""
    
    # Reemplazar caracteres inválidos + comillas Unicode
    base = re.sub(
        r'[\\/:*?"<>|\r\n\t“”‘’´`]',
        "_",
        base
    )
    
    # Normalizar espacios
    base = re.sub(r'\s+', " ", base)
    base = base.strip()

    # Limitar longitud sin afectar extensión
    if len(base) > 70:
        base = base[:70]

    # Volver a unir con extensión (en minúsculas)
    return f"{base}{ext.lower()}"

def quitar_acentos(texto):
    reemplazos = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U",
        "ñ": "n", "Ñ": "N"
    }
    for a, b in reemplazos.items():
        texto = texto.replace(a, b)
    return texto

def nombres_conocidos_cc(cc):
    if not cc:
        return "-----"
    autorizados = [
        "Arq. Belty Espinoza", "Arq. Ricardo Valverde", "Ing. Jhon Araujo", "Arq. Ader Intriago",
        "Ing. Jorge Maldonado", "Arq. Roberto Vivanco", "Ing. Lissette Moreno",
        "Ing. Luis Vargas", "Ing. Patricia Fuentes", "Lic. Narcisa Munoz", "Lic. Leonardo Rodriguez",
        "Ing. Jaime Franco", "Arq. Jhonther Cardenas", "Ing. Miguel Flores", "Ing. Jorge Tohaquiza",
        "Ing. Humberto Rodriguez", "Ing. Eddy Alfonso", "Sra. Gilda Suarez", "Arq. Pilar Zalamea",
        "Ing. Isaac Munoz", "Arq. Franklin Medina", "Ing. Victor Velasco", "Sra. Bethzaida Villamil"
    ]
    cc_norm = quitar_acentos(cc.lower())
    resultado = []
    for nombre in autorizados:
        n_norm = quitar_acentos(nombre.lower())
        partes = n_norm.split()
        if len(partes) >= 2 and partes[1] in cc_norm and partes[-1] in cc_norm:
            resultado.append(nombre)
    if len(resultado) == 0:
        return "-----"
    else:
        return "\n".join(resultado)

def nombres_conocidos_rem(rem):
    if not rem:
        return ""
    autorizados = [
        "Arq. Belty Espinoza Santos", "Arq. Ricardo Valverde Paredes", "Ing. Jhon Araujo Borjas", "Arq. Ader Intriago Arteaga",
        "Ing. Jorge Maldonado Granizo", "Arq. Roberto Vivanco Calderon", "Ing. Lissette Moreno Balladares",
        "Ing. Luis Vargas Orozco", "Ing. Patricia Fuentes Moran", "Lic. Narcisa Munoz Feraud", "Lic. Leonardo Rodriguez Molina",
        "Ing. Jaime Franco Baquerizo", "Arq. Jhonther Cardenas", "Ing. Miguel Flores Poveda", "Ing. Jorge Tohaquiza Jacho",
        "Ing. Humberto Rodriguez Gonzalez", "Ing. Eddy Alfonso Garcia", "Sra. Gilda Suarez Crespin", "Arq. Pilar Zalamea Garcia",
        "Ing. Isaac Munoz Mindiola", "Arq. Franklin Medina Gonzalez", "Ing. Victor Velasco Matute", "Sra. Bethzaida Villamil",
        "Ab. María Marroquín Mora", "Ing. Jorge Gutiérrez Tenorio", "Mgs. Jhair Jiménez Aldaz"
    ]
    rem_norm = quitar_acentos(rem.lower())
    resultado = []
    for nombre in autorizados:
        n_norm = quitar_acentos(nombre.lower())
        partes = n_norm.split()
        if len(partes) >= 2 and partes[1] in rem_norm and partes[2] in rem_norm:
            resultado.append(nombre)
    
    if len(resultado) == 0:
        return rem
    else:
        return "\n".join(resultado)

def cut_nombres_destinatarios(destinatario: str) -> str:
    if not destinatario:
        return ""
    
    # Separar por ; o ,
    personas = re.split(r'[;,]', destinatario)
    resultado = []

    for p in personas:
        p = p.strip()
        if not p:
            continue

        # Quitar el correo dentro de <>
        p = re.sub(r"<.*?>", "", p).strip()

        partes = p.split()
        
        if len(partes) == 1:
            # Solo una palabra → se deja igual
            resultado.append(partes[0])
        elif len(partes) == 2:
            # nombre + apellido
            nombre = partes[0]
            apellido = partes[1]
            resultado.append(f"{nombre} {apellido}")
        else:
            # 3 o más → penúltima palabra es primer apellido
            nombre = partes[0]
            apellido = partes[-2]
            resultado.append(f"{nombre} {apellido}")

        # # Si vienen 2 nombres + 2 apellidos
        # if len(partes) == 4:
        #     nombre = partes[0]       # primer nombre
        #     apellido = partes[2]     # primer apellido
        #     resultado.append(f"{nombre} {apellido}")
        # else:
        #     # fallback: tomar solo la primera palabra
        #     resultado.append(partes[0])

    # Unir con salto de línea
    # return "\n".join(resultado)
    return resultado

def obtener_info_destinatarios(lista_nombres):
    destinatarios = []
    cargos = []
    for nombre_abreviado in lista_nombres:
        n = nompropio_python(nombre_abreviado)
        persona = DESTINATARIOS_CONOCIDOS.get(n)
        if persona:
            destinatarios.append(persona["nombre_completo"])
            cargos.append(persona["cargo"])
        else:
            destinatarios.append(nombre_abreviado)
            # cargos.append("")
            
    return ("\n\n".join(destinatarios)), ("\n\n".join(cargos))

def obtener_info_remitente(remitente):
    nombre_limpio = limpiar_nombre(remitente)
    palabras = nombre_limpio.split()

    for nombre_base, info in REMITENTES_LIMPIO.items():
        coincidencias = 0
        for p in palabras:
            if p in nombre_base:
                coincidencias += 1

        # Se considera coincidencia válida si al menos 2 palabras coinciden
        if coincidencias >= 2:
            return info["cargo"], info["dependencia"]

    return None, None

def limpiar_nombre(nombre):
    if not nombre:
        return ""

    # minúsculas
    nombre = nombre.lower()

    # quitar tildes
    nombre = ''.join(
        c for c in unicodedata.normalize('NFD', nombre)
        if unicodedata.category(c) != 'Mn'
    )

    # quitar títulos comunes
    nombre = re.sub(r'\b(arq|ing|lic|sr|sra|dra|dr|ab)\.?', '', nombre)

    # quitar puntos y comas
    nombre = nombre.replace('.', '').replace(',', '')

    # quitar espacios dobles
    nombre = re.sub(r'\s+', ' ', nombre).strip()

    return nombre

def nompropio_python(asunto):
    return asunto.title()

def exportar_excel(registros, carpeta_base, fecha_inicio_str):
    # === Exportar resultados a Excel con formato ===
    if registros:
        df = pd.DataFrame(registros)
        
        ruta_excel = carpeta_base / f"{fecha_inicio_str}_CorreosExportados.xlsx"
        
        df.to_excel(ruta_excel, index=False)

        # === Aplicar formato con openpyxl ===
        wb = load_workbook(ruta_excel)
        ws = wb.active

        # Cuerpo
        for row in ws.iter_rows():
            for celda in row:
                celda.font = Font(name="Arial", size=12)
                # if celda.value is not None:   # opcional: solo celdas con contenido
                celda.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

        for celda in ws["I"]:
            celda.font = Font(name="Arial", size=12, color="FF0000", bold=True)  # rojo
            celda.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Bordes
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Aplicar bordes a las celdas
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # Ajustar ancho de columnas
        for col in ws.columns:
            column = get_column_letter(col[0].column)
            ws.column_dimensions[column].width = 20.71
        
        ws.column_dimensions["G"].width = 30.71
        ws.column_dimensions["I"].width = 30.71
        
        # Ajustar alto de filas
        for row in ws.iter_rows():
            celda = row[0]
            ws.row_dimensions[celda.row+1].height = 150.04

        # Convertir a tabla con estilo
        ultima_fila = ws.max_row
        ultima_col = ws.max_column
        rango_tabla = f"A1:{get_column_letter(ultima_col)}{ultima_fila}"

        # Llenar con color amarillo las celdas vacías
        body_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") #amarillo
        
        for fila in ws[rango_tabla]:     # fila = tuple de celdas
            for celda in fila:           # celda = objeto real de openpyxl
                if celda.value is None or celda.value == "":
                    celda.fill = body_fill

        # Tabla
        tabla = Table(displayName="Exportados", ref=rango_tabla)
        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo
        ws.add_table(tabla)
        ws.sheet_view.zoomScale = 80 # Ajustar zoom
        wb.save(ruta_excel)

        print(Fore.GREEN + f"\n\nExportación completada correctamente: {len(registros)} registros.")
        input("\nPresione ENTER para abrir la carpeta.")
        os.system("cls")
        os.startfile(carpeta_base)
    else:
        print(Fore.RED + "\n\nNo se encontraron Correos en el rango especificado.")
        input("\nPresione ENTER para continuar.")
        os.system("cls")    

def obtener_anexos(anexos, carpeta_correo):
    lista_anexos = []

    if anexos.Count > 0:
        # i = 0
        for anexo in anexos:
            nombre_archivo = limpiar_texto(anexo.FileName)
            if "image" not in nombre_archivo.lower() and "outlook" not in nombre_archivo.lower():
                carpeta_anexos = carpeta_correo / "Anexos"
                os.makedirs(carpeta_anexos, exist_ok=True) # crear carpeta de anexos   
                ruta_anexo = carpeta_anexos / nombre_archivo
                anexo.SaveAsFile(str(ruta_anexo))
                # i+=1
                lista_anexos.append(nombre_archivo)
    return lista_anexos

def pedir_fecha():
    dia = input("Día (1-31): ")
    mes = input("Mes (1-12): ")
    anio = input("Año (4 dígitos): ")
    
    # Fechas en formato datetime
    fecha_inicio_raw = datetime(int(anio), int(mes), int(dia), 0, 0, 0)
    fecha_fin_raw = fecha_inicio_raw + timedelta(days=1)

    # Fecha inicio en string
    fecha_inicio_str = fecha_inicio_raw.strftime("%Y-%m-%d")

    # Obtener nombre del mes
    meses = {1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril", 5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto", 9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"}
    mes_int= int(mes)
    mes_name = meses[mes_int]
    
    return fecha_inicio_raw, fecha_fin_raw, fecha_inicio_str, mes_name, anio

def procesar():
    print("=== Exportador de Correos y Anexos ===", end="\n\n")
        
    # Inicializa colorama
    init(autoreset=True)  # Después de cada print el color vuelva al normal

    # Pedir fechas al usuario
    f_inicio, f_fin, f_inicio_str, mes_name, anio = pedir_fecha()
    
    # Ruta carpeta base "Correos Agosto - 2025 / 2025-08-30"
    carpeta_base = Path.home() / "Downloads" / f"Correos {mes_name} - {anio}" / f_inicio_str
    
    # Conectar con Outlook
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    carpeta = outlook.GetDefaultFolder(6)  # 6 = Bandeja de entrada

    # Filtro para Outlook
    filtro = f"[ReceivedTime] >= '{f_inicio.strftime('%d/%m/%Y %H:%M')}' AND [ReceivedTime] < '{f_fin.strftime('%d/%m/%Y %H:%M')}'"
    
    # Obtener correos filtrados
    items = carpeta.Items
    items.Sort("[ReceivedTime]", False)   # True = descendente, False = ascendente
    items_filtrados = items.Restrict(filtro)
    
    # Abrir Word
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    
    # Mostrar "Procesando"
    if len(items_filtrados) > 0:
        print("\nProcesando", end="")
    
    # Array de registros
    registros = []
    
    # Pausa corta
    time.sleep(0.5)

    for mail in items_filtrados:
        try:
            # Si no es correo, omitir
            if mail.Class != 43:
                continue
            
            # Setear info del correo
            recibido = mail.ReceivedTime
            remitente = mail.SenderName
            anexos = mail.Attachments
            asunto = mail.Subject or ""
            destinatarios_raw = mail.To
            cc = mail.CC or ""
            
            # Fecha python
            recibido_py = datetime(recibido.year, recibido.month, recibido.day,
                                recibido.hour, recibido.minute, recibido.second)

            # Si no está dentro del rango, omitir
            if not (f_inicio <= recibido_py < f_fin):
                continue
            
            # Si contiene esto en remitente, omitir
            if remitente in ["QUIPUX", "UNIVERSIDAD DE GUAYAQUIL", "INFO UG", "Comunicados DVSBE", "Zoom", "Titulares EL UNIVERSO", "Canva", "ClickUp Notifications", "ClickUp Team", "DepositPhotos"]:
                continue
            
            # . . .
            print(".", end="", flush=True)
            
            # Crear carpeta del correo
            # id_correo = recibido_py.strftime("%H%M%S") + "_" + limpiar_texto(remitente)
            id_correo = recibido_py.strftime("%H%M%S") + "_" + limpiar_acortar_remitentes(remitente)
            carpeta_nombre = f"{id_correo}"
            
            # Acortar nombre de carpeta
            if len(carpeta_nombre) > 100:
                carpeta_nombre = carpeta_nombre[:100]

            # Crear carpeta del día
            carpeta_base.mkdir(parents=True, exist_ok=True)
            carpeta_correo = carpeta_base / carpeta_nombre
            os.makedirs(carpeta_correo, exist_ok=True)

            asunto_limpio = limpiar_texto(asunto)

            # Rutas para guardar correo
            mht_path = carpeta_correo / f"{asunto_limpio}.mht"
            pdf_path = carpeta_correo / f"{asunto_limpio}.pdf"
            msg_path = carpeta_correo / f"{asunto_limpio}.msg"

            # Conversión del correo a .pdf
            try:
                mail.SaveAs(str(mht_path), 10)
                doc = word.Documents.Open(str(mht_path))
                
                 # Ajustar tamaño de todas las imágenes (max ancho 500 pts aprox)
                for shape in doc.InlineShapes:
                    if shape.Type in [1, 3, 4, 5]:  # Tipos de imagen: 1=Embedded, 3=Linked, etc.
                        max_width = 800  # puedes ajustar
                        if shape.Width > max_width:
                            ratio = max_width / shape.Width
                            shape.Width = max_width
                            shape.Height = shape.Height * ratio
                
                doc.ExportAsFixedFormat(OutputFileName=str(pdf_path), ExportFormat=17)
                doc.Close(False)
                os.remove(mht_path)
            except Exception:
                mail.SaveAs(str(msg_path), 3)
            
            # Agregar título a remitente conocido
            remitente_filtrado = nombres_conocidos_rem(remitente)
            
            # Obtener cargo y facultad del remitente conocido
            cargo, dependencia = obtener_info_remitente(remitente)
            
            # Obtener lista de destinatarios con saltos de línea
            destinatarios_cortos = cut_nombres_destinatarios(destinatarios_raw)
            # destinatarios_final = aplicar_titulos(destinatarios_cortos)
            destinatarios_final, cargos = obtener_info_destinatarios(destinatarios_cortos)
            
            # Filtrar nombres conocidos en CC
            cc_filtrado = nombres_conocidos_cc(str(cc))
            
            # Guardar Anexos
            lista_anexos = obtener_anexos(anexos, carpeta_correo)
            
            # Obtener cantidad de anexos
            cant_anexos = len(lista_anexos)
            
            # Mostrar cantidad de anexos
            observaciones = "No contiene anexos" if cant_anexos == 0 else f"Anexa {cant_anexos} documento(s)"

            # Guardar registros
            registros.append({
                "Fecha del Documento": recibido_py.strftime("%Y-%m-%d %H:%M:%S"),
                "Remitente": nompropio_python(remitente_filtrado),
                "Cargo": cargo,
                "Facultad/Dependencia": dependencia,
                "Destinatario": nompropio_python(destinatarios_final),
                # "Destinatario": nompropio_python(destinatarios_formatted),
                "Empresa/Cargo": cargos,
                "Asunto": nompropio_python(asunto),
                "Con Copia": cc_filtrado,
                "Observaciones": observaciones
            })
        except Exception as e:
            print(f"Error procesando correo: {e}")

    # Cerrar Word
    word.Quit()
    
    # Crear y exportar excel con registros
    exportar_excel(registros, carpeta_base, f_inicio_str)
 
REMITENTES_CONOCIDOS = {
    # Dirección DIOU
    "Arq. Belty Espinoza Santos": {
        "cargo": "Directora",
        "dependencia": "Dirección DIOU"
    },

    # Jefatura de Planificación
    "Arq. Ricardo Valverde Paredes": {
        "cargo": "Jefe de Planificación",
        "dependencia": "Jefatura de Planificación"
    },
    "Ing. Jhon Araujo Borjas": {
        "cargo": "Analista de Planificación de Obras 3",
        "dependencia": "Jefatura de Planificación"
    },
    "Arq. Ader Intriago Arteaga": {
        "cargo": "Analista de Planificación de Obras 3",
        "dependencia": "Jefatura de Planificación"
    },
    "Ing. Jorge Maldonado Granizo": {
        "cargo": "Analista de Planificación de Obras 3",
        "dependencia": "Jefatura de Planificación"
    },
    "Arq. Roberto Vivanco Calderón": {
        "cargo": "Analista de Planificación de Obras 3",
        "dependencia": "Jefatura de Planificación"
    },
    "Ing. Lissette Moreno Balladares": {
        "cargo": "Analista de Planificación de Obras 3",
        "dependencia": "Jefatura de Planificación"
    },
    "Ing. Luis Vargas Orozco": {
        "cargo": "Analista de Planificación de Obras 3",
        "dependencia": "Jefatura de Planificación"
    },
    "Ing. Patricia Fuentes Morán": {
        "cargo": "Analista de Planificación 1",
        "dependencia": "Jefatura de Planificación"
    },
    "Lic. Narcisa Muñoz Feraud": {
        "cargo": "Asistente de Planificación de Obras 2",
        "dependencia": "Jefatura de Planificación"
    },
    "Lic. Leonardo Rodríguez Molina": {
        "cargo": "Asistente de Planificación de Obras 2",
        "dependencia": "Jefatura de Planificación"
    },
    "Ing. Jaime Franco Baquerizo": {
        "cargo": "Asistente 2 de Infraestructura y Obras Universitarias",
        "dependencia": "Jefatura de Planificación"
    },
    "Arq. Jhonther Cárdenas": {
        "cargo": "Supervisor de Construcciones Civiles",
        "dependencia": "Jefatura de Planificación"
    },

    # Jefatura de Infraestructura
    "Ing. Miguel Flores Poveda": {
        "cargo": "Jefe de Infraestructura",
        "dependencia": "Jefatura de Infraestructura"
    },
    "Ing. Jorge Tohaquiza Jacho": {
        "cargo": "Analista de Infraestructura de Obras Universitarias 3",
        "dependencia": "Jefatura de Infraestructura"
    },
    "Ing. Humberto Rodríguez González": {
        "cargo": "Analista de Mantenimiento de Infraestructura 3",
        "dependencia": "Jefatura de Infraestructura"
    },
    "Ing. Eddy Alfonso García": {
        "cargo": "Asistente de Infraestructura de Obras Universitarias 2",
        "dependencia": "Jefatura de Infraestructura"
    },
    "Gilda Suárez Crespín": {
        "cargo": "Asistente de Infraestructura de Obras Universitarias 2",
        "dependencia": "Jefatura de Infraestructura"
    },

    # Jefatura de Diseño y Fiscalización
    "Arq. Pilar Zalamea García": {
        "cargo": "Jefe de Diseño y Fiscalización",
        "dependencia": "Jefatura de Diseño y Fiscalización"
    },
    "Ing. Isaac Muñoz Mindiola": {
        "cargo": "Analista de Diseño y Fiscalización 1",
        "dependencia": "Jefatura de Diseño y Fiscalización"
    },
    "Arq. Franklin Medina González": {
        "cargo": "Analista de Diseño y Fiscalización 3",
        "dependencia": "Jefatura de Diseño y Fiscalización"
    },
    "Ing. Victor Velasco Matute": {
        "cargo": "Asistente de Infraestructura 2",
        "dependencia": "Jefatura de Diseño y Fiscalización"
    },
    "Sr. Bethzaida Villamil": {
        "cargo": "Asistente 2 Diseño y Planificación",
        "dependencia": "Jefatura de Diseño y Fiscalización"
    },
    
    #Otros
    "Ab. María Marroquín Mora": {
        "cargo": "Analista de Talento Humano",
        "dependencia": "DIOU"
    },
    "Ing. Jorge Gutiérrez Tenorio":{
        "cargo": "Contratista",
        "dependencia": "JKGT"
    },
    "Mgs. Jhair Jiménez Aldaz":{
        "cargo": "Director de Compras Públicas",
        "dependencia": ""
    }
}

REMITENTES_LIMPIO = {
    limpiar_nombre(nombre_original): info
    for nombre_original, info in REMITENTES_CONOCIDOS.items()
}

DESTINATARIOS_CONOCIDOS = {
    "Belty Espinoza": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Belty Espinoza",
        "cargo": "Directora DIOU"
    },
    "Ricardo Valverde": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Ricardo Valverde",
        "cargo": "Jefe de Planificación"
    },
    "Jhon Araujo": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Jhon Araujo",
        "cargo": "Analista de Planificación de Obras 3"
    },
    "Ader Intriago": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Ader Intriago",
        "cargo": "Analista de Planificación de Obras 3"
    },
    "Jorge Maldonado": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Jorge Maldonado",
        "cargo": "Analista de Planificación de Obras 3"
    },
    "Roberto Vivanco": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Roberto Vivanco",
        "cargo": "Analista de Planificación de Obras 3"
    },
    "Lissette Moreno": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Lissette Moreno",
        "cargo": "Analista de Planificación de Obras 3"
    },
    "Luis Vargas": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Luis Vargas",
        "cargo": "Analista de Planificación de Obras 3"
    },
    "Patricia Fuentes": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Patricia Fuentes",
        "cargo": "Analista de Planificación 1"
    },
    "Narcisa Munoz": {
        "titulo": "Lic.",
        "nombre_completo": "Lic. Narcisa Munoz",
        "cargo": "Asistente de Planificación de Obras 2"
    },
    "Leonardo Rodriguez": {
        "titulo": "Lic.",
        "nombre_completo": "Lic. Leonardo Rodriguez",
        "cargo": "Asistente de Planificación de Obras 2"
    },
    "Jaime Franco": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Jaime Franco",
        "cargo": "Asistente 2 de Infraestructura y Obras Universitarias"
    },
    "Jhonther Cardenas": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Jhonther Cardenas",
        "cargo": "Supervisor de Construcciones Civiles"
    },
    "Miguel Flores": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Miguel Flores",
        "cargo": "Jefe de Infraestructura"
    },
    "Jorge Tohaquiza": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Jorge Tohaquiza",
        "cargo": "Analista de Infraestructura de Obras Universitarias 3"
    },
    "Humberto Rodriguez": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Humberto Rodriguez",
        "cargo": "Analista de Mantenimiento de Infraestructura 3"
    },
    "Eddy Alfonso": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Eddy Alfonso",
        "cargo": "Asistente de Infraestructura de Obras Universitarias 2"
    },
    "Gilda Suarez": {
        "titulo": "Sra.",
        "nombre_completo": "Sra. Gilda Suarez",
        "cargo": "Asistente de Infraestructura de Obras Universitarias 2"
    },
    "Pilar Zalamea": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Pilar Zalamea",
        "cargo": "Jefe de Diseño y Fiscalización"
    },
    "Isaac Munoz": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Isaac Munoz",
        "cargo": "Analista de Diseño y Fiscalización 1"
    },
    "Franklin Medina": {
        "titulo": "Arq.",
        "nombre_completo": "Arq. Franklin Medina",
        "cargo": "Analista de Diseño y Fiscalización 3"
    },
    "Victor Velasco": {
        "titulo": "Ing.",
        "nombre_completo": "Ing. Victor Velasco",
        "cargo": "Asistente de Infraestructura 2"
    },
    "Bethzaida Villamil": {
        "titulo": "Sra.",
        "nombre_completo": "Sra. Bethzaida Villamil",
        "cargo": "Asistente 2 Diseño y Planificación"
    },
    
    # Otros
    "Jhair Jimenez":{
        "titulo": "Mgs.",
        "nombre_completo": "Mgs. Jhair Jimenez",
        "cargo": "Director de Compras Públicas"
    }
}

if __name__ == "__main__":
    while True:
        procesar()